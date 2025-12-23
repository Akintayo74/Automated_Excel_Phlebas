"""
Student Portal Automation Script
Automatically fetches admission numbers from the school portal and updates Excel file
"""

import time
import re
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import openpyxl
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher
from dotenv import load_dotenv
import os
import argparse
import glob
from pathlib import Path
import logging
from datetime import datetime

load_dotenv()
# Add this helper function at the top of your script
def get_similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()

class StudentPortalScraper:
    def __init__(self, excel_path, portal_url):
        self.excel_path = excel_path
        self.portal_url = portal_url
        self.driver = None
        self.wb = None
        self.ws = None
        
    def setup_logging(self, log_dir="./logs"):
        """Setup file logging for the scraping session"""
        import os
        from pathlib import Path
        
        # Create logs directory if it doesn't exist
        Path(log_dir).mkdir(exist_ok=True)
        
        # Create log filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_name = Path(self.excel_path).stem
        log_file = f"{log_dir}/{excel_name}_{timestamp}.log"
        
        # Setup logger
        self.logger = logging.getLogger(f"Scraper_{excel_name}")
        self.logger.setLevel(logging.INFO)
        
        # File handler
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        
        # Console handler (still print important stuff)
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.WARNING)  # Only warnings/errors to console
        
        # Format
        formatter = logging.Formatter(
            '%(asctime)s | %(levelname)s | %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)
        
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        
        self.logger.info("="*80)
        self.logger.info(f"SCRAPING SESSION STARTED: {excel_name}")
        self.logger.info(f"Log file: {log_file}")
        self.logger.info("="*80)
        
        print(f"📝 Logging to: {log_file}")
        return log_file

    def setup_driver(self):
        """Initialize Chrome WebDriver with appropriate options"""
        chrome_options = Options()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        # Remove headless mode so you can see what's happening
        # chrome_options.add_argument('--headless')
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.maximize_window()
        print("✓ Browser initialized")
        
    def login(self, username, password):
        """Login to the portal"""
        try:
            print(f"\n→ Navigating to {self.portal_url}")
            self.driver.get(self.portal_url)
            
            # Wait for login page to load
            print("→ Waiting for login page...")
            time.sleep(3)
            
            # You'll need to inspect the login page to find the correct selectors
            # These are common field names - adjust if needed
            print("→ Attempting to log in...")
            
            # Try to find username/email field
            try:
                username_field = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.NAME, "email"))
                )
            except:
                username_field = self.driver.find_element(By.NAME, "username")
            
            username_field.clear()
            username_field.send_keys(username)
            
            # Find password field
            password_field = self.driver.find_element(By.NAME, "password")
            password_field.clear()
            password_field.send_keys(password)
            
            # Find and click login button
            login_button = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            login_button.click()
            
            # Wait for navigation after login
            print("→ Logging in...")
            time.sleep(5)
            
            # Check if login was successful by looking for student list page
            if "students" in self.driver.current_url.lower() or "dashboard" in self.driver.current_url.lower():
                print("✓ Login successful!")
                return True
            else:
                print("✗ Login may have failed. Please check credentials.")
                return False
                
        except Exception as e:
            print(f"✗ Login error: {str(e)}")
            return False
    
    def clean_name(self, name):
        """Clean student name for searching"""
        if not name or name == "NAME":
            return None
        # Remove extra spaces and convert to title case
        name = re.sub(r'\s+', ' ', str(name).strip())
        # Extract last name (usually the first word in your format)
        parts = name.split()
        if len(parts) > 0:
            return parts[0]  # Return first word (last name)
        return name
    
    def search_student(self, name, full_name):
        try:
            if "students" not in self.driver.current_url.lower():
                self.driver.get(self.portal_url)
                time.sleep(2)

             # 1. Set Class Filter if specified
            if hasattr(self, 'target_class') and self.target_class:
                try:
                    # Find the CLASS dropdown
                    class_dropdown = self.driver.find_element(
                        By.XPATH, 
                        "//label[contains(text(), 'CLASS')]/following::select[1]"
                    )
                    
                    # Use Select to choose the class
                    select = Select(class_dropdown)
                    
                    # Try exact match first
                    try:
                        select.select_by_visible_text(self.target_class)
                        print(f"  ✓ Class filter set to: {self.target_class}")
                        time.sleep(1)
                    except:
                        # Try partial match if exact fails
                        for option in select.options:
                            if self.target_class.upper() in option.text.upper():
                                select.select_by_visible_text(option.text)
                                print(f"  ✓ Class filter set to: {option.text}")
                                time.sleep(1)
                                break
                except Exception as e:
                    print(f"  ⚠ Could not set class filter: {e}")

            search_box = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='text']"))
            )
            search_box.clear()
            search_box.send_keys(name)
            time.sleep(2.5) # Increased wait for portal to refresh results

            rows = self.driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            
            full_name_clean = re.sub(r'\s+', ' ', full_name.upper().strip())
            full_name_parts = [p for p in full_name_clean.split() if len(p) >= 3]

            best_match = None
            best_score = 0

            for row in rows:
                try:
                    # FIX 1: Verify this is a data row, not a "No Results" message
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) < 3: 
                        continue

                    admission_number = cells[0].text.strip()
                    first_name_portal = cells[1].text.strip().upper()
                    last_name_portal = cells[2].text.strip().upper()
                    
                    portal_display = f"{first_name_portal} {last_name_portal}"
                    portal_full_no_space = portal_display.replace(" ", "")
                    portal_words = set(portal_display.split())

                    # Method 1: Exact word matching
                    matching_words = set(full_name_parts).intersection(portal_words)
                    exact_score = len(matching_words) / len(full_name_parts) if full_name_parts else 0

                    # FIX 2: Improved Fuzzy Matching with SequenceMatcher
                    fuzzy_points = 0
                    for part in full_name_parts:
                        if part in portal_full_no_space:
                            fuzzy_points += 1
                        else:
                            # Check similarity against each word in the portal name
                            word_sims = [get_similarity(part, w) for w in portal_words]
                            max_sim = max(word_sims) if word_sims else 0
                            if max_sim > 0.8: # Threshold for spelling variants
                                fuzzy_points += max_sim
                    
                    fuzzy_score = fuzzy_points / len(full_name_parts) if full_name_parts else 0
                    normalized_score = max(exact_score, fuzzy_score)

                    print(f" • {portal_display}: {admission_number} (Score: {normalized_score:.0%})")

                    if normalized_score > best_score:
                        best_score = normalized_score
                        best_match = (admission_number, portal_display)

                except Exception as row_err:
                    # Useful for debugging why a row failed
                    # print(f"  ⚠ Row error: {row_err}")
                    continue

            if best_match and best_score >= 0.45:
                # Log the match with confidence score
                if hasattr(self, 'logger'):
                    if best_score < 0.70:
                        self.logger.warning(
                            f"LOW CONFIDENCE | Row {getattr(self, 'current_row', '?')} | "
                            f"Excel: {full_name} | Portal: {best_match[1]} | "
                            f"Admission: {best_match[0]} | Score: {best_score:.2%}"
                        )
                    else:
                        self.logger.info(
                            f"MATCHED | Row {getattr(self, 'current_row', '?')} | "
                            f"Excel: {full_name} | Portal: {best_match[1]} | "
                            f"Admission: {best_match[0]} | Score: {best_score:.2%}"
                        )
                
                # Print to console
                if best_score < 0.70:
                    print(f" ⚠ LOW CONFIDENCE MATCH ({best_score:.0%}) - Please verify!")
                print(f" ✓ Best match: {best_match[1]} → {best_match[0]}")
                return best_match[0]
            else:
                # Log failed matches
                if hasattr(self, 'logger'):
                    self.logger.error(
                        f"NO MATCH | Row {getattr(self, 'current_row', '?')} | "
                        f"Excel: {full_name} | Best score: {best_score:.2%} | "
                        f"Searched: {name}"
                    )
                return None

        except Exception as e:
            print(f" ✗ Search error: {str(e)}")
            return None
    
    def load_excel(self):
        """Load the Excel file"""
        try:
            self.wb = openpyxl.load_workbook(self.excel_path)
            self.ws = self.wb.active
            print(f"✓ Excel file loaded: {self.excel_path}")
            return True
        except Exception as e:
            print(f"✗ Error loading Excel: {str(e)}")
            return False
        
    def process_workbook(self, start_row=3):
        """NEW: This handles the loop through all sheets (Silver, Gold, Ruby)"""
        total_updated = 0
        total_skipped = 0
        total_errors = 0
        
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Loop through all sheets in the workbook
        for sheet_name in self.wb.sheetnames:
            print(f"\n{'='*70}")
            print(f"  PROCESSING SHEET: {sheet_name}")
            print(f"{'='*70}")
            
            # Set the current active sheet
            self.ws = self.wb[sheet_name]
            
            # Call the row-by-row processing logic
            updated, skipped, errors = self.process_students(start_row, yellow_fill)
            
            total_updated += updated
            total_skipped += skipped
            total_errors += errors
        
        # Save the final result after ALL sheets are done
        try:
            output_path = self.excel_path.replace('.xlsx', '_updated.xlsx')
            self.wb.save(output_path)
            
            # Log final summary
            if hasattr(self, 'logger'):
                self.logger.info("="*80)
                self.logger.info(f"WORKBOOK SAVED: {output_path}")
                self.logger.info(f"Total Updated: {total_updated} | Skipped: {total_skipped} | Errors: {total_errors}")
                self.logger.info("="*80)
            
            print(f"\n{'='*70}")
            print(f"✓ ENTIRE WORKBOOK SAVED: {output_path}")
            print(f"{'='*70}")
            print(f"\nFINAL SUMMARY (All Sheets):")
            print(f"  • Total Updated: {total_updated}")
            print(f"  • Total Skipped: {total_skipped}")
            print(f"  • Total Errors: {total_errors}")
            print(f"  • Total Processed: {total_updated + total_skipped + total_errors}")
            return output_path, total_updated, total_errors
        except Exception as e:
            print(f"\n✗ Error saving workbook: {str(e)}")
            if hasattr(self, 'logger'):
                self.logger.error(f"Error saving workbook: {str(e)}")
            return None, total_updated, total_errors
    
    def process_students(self, start_row=3, yellow_fill=None):
        """MODIFIED: Now returns counts instead of saving the file here"""
        if not self.ws:
            print("✗ Excel file not loaded")
            return 0, 0, 0
        
        # Create yellow fill if not provided
        if yellow_fill is None:
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        updated_count = 0
        skipped_count = 0
        error_count = 0
        
        for row_idx in range(start_row, self.ws.max_row + 1):
            admission_cell = self.ws.cell(row=row_idx, column=1)  # Column A
            name_cell = self.ws.cell(row=row_idx, column=2)        # Column B
            
            current_admission = admission_cell.value
            student_name = name_cell.value
            
            # Skip if no name
            if not student_name or student_name == "NAME":
                continue
            
            print(f"\nRow {row_idx}: {student_name}")
            
            # Skip if admission number already exists and looks valid
            if current_admission and isinstance(current_admission, str) and "CDSSJOS" in current_admission:
                print(f"  ⊘ Skipped (already has admission number: {current_admission})")
                skipped_count += 1
                continue
            
            # Clean and search for the student
            search_name = self.clean_name(student_name)
            if not search_name:
                print(f"  ⊘ Skipped (couldn't parse name)")
                skipped_count += 1
                continue

            # Track current row for logging
            self.current_row = row_idx 
            
            # Search and get admission number
            admission_number = self.search_student(search_name, student_name)
            
            if admission_number:
                # Update Excel
                admission_cell.value = admission_number
                admission_cell.fill = yellow_fill  # Highlight updated cells
                updated_count += 1
                print(f"  ✓ Updated in Excel")
            else:
                error_count += 1
            
            # Small delay to avoid overwhelming the server
            time.sleep(1)
        
        # Print sheet summary
        print(f"\n{'-'*70}")
        print(f"Sheet Summary ({self.ws.title}):")
        print(f"  • Updated: {updated_count}")
        print(f"  • Skipped: {skipped_count}")
        print(f"  • Errors: {error_count}")
        print(f"{'-'*70}")
        
        return updated_count, skipped_count, error_count

    def close(self):
        """Close browser and cleanup"""
        if self.driver:
            self.driver.quit()
            print("\n✓ Browser closed")


def main():
    """Main execution function with command-line argument support"""
    parser = argparse.ArgumentParser(
        description="Automated Student Portal Scraper",
        epilog="Example: python script.py /path/to/folder/ --class 'JSS 3'"
    )
    parser.add_argument(
        "path", 
        help="Path to Excel file(s) or directory containing Excel files"
    )
    parser.add_argument(
        "--pattern",
        default="*.xlsx",
        help="File pattern to match (default: *.xlsx)"
    )
    parser.add_argument(
        "--class",
        dest="student_class",
        default=None,
        help="Filter by class (e.g., 'JSS 3', 'SS 3') - speeds up search"
    )
    args = parser.parse_args()

    # 2. Get Credentials from .env
    PORTAL_URL = os.getenv("PORTAL_URL")
    USERNAME = os.getenv("PORTAL_USER")
    PASSWORD = os.getenv("PORTAL_PASS")
    
    # Validate environment variables
    if not all([PORTAL_URL, USERNAME, PASSWORD]):
        print("✗ Error: Missing environment variables!")
        print("  Please ensure your .env file contains:")
        print("    PORTAL_URL=https://...")
        print("    PORTAL_USER=your_username")
        print("    PORTAL_PASS=your_password")
        return

    # 3. Determine which files to process
    path = Path(args.path)
    
    if path.is_file():
        # Single file provided
        files_to_process = [str(path)]
    elif path.is_dir():
        # Directory provided - find all Excel files
        pattern = args.pattern
        files_to_process = glob.glob(str(path / pattern))
        if not files_to_process:
            print(f"✗ No files matching '{pattern}' found in {path}")
            return
    else:
        print(f"✗ Path does not exist: {path}")
        return
    
    # 4. Display files to process
    print("\n" + "="*70)
    print("STUDENT PORTAL AUTOMATION SCRIPT")
    print("="*70)
    print(f"Files to process: {len(files_to_process)}")
    for f in files_to_process:
        print(f"  • {Path(f).name}")
    print("="*70 + "\n")
    
    # 5. Process each file
    
    for idx, file_path in enumerate(files_to_process, 1):
        print(f"\n{'#'*70}")
        print(f">>> FILE {idx}/{len(files_to_process)}: {file_path}")
        print(f"{'#'*70}")
        
        scraper = StudentPortalScraper(file_path, PORTAL_URL)
        scraper.target_class = args.student_class
        scraper.setup_logging()
        
        try:
            scraper.setup_driver()
            
            if not scraper.login(USERNAME, PASSWORD):
                print(f"\n✗ Failed to login for {file_path}. Skipping...")
                scraper.close()
                continue
            
            if not scraper.load_excel():
                print(f"\n✗ Failed to load {file_path}. Skipping...")
                scraper.close()
                continue
            
            # Process all sheets in the workbook
            output, updated, errors = scraper.process_workbook(start_row=3)
            
            if output:
                print(f"\n✓ FINISHED: {file_path}")
                print(f"✓ Saved to: {output}")
                print(f"✓ Total Updated: {updated} | Total Errors: {errors}")
            
        except KeyboardInterrupt:
            print("\n\n⚠ Process interrupted by user")
            scraper.close()
            break
        except Exception as e:
            print(f"✗ Fatal error processing {file_path}: {e}")
        finally:
            scraper.close()
    
    print(f"\n{'='*70}")
    print("ALL FILES PROCESSED")
    print(f"{'='*70}\n")

if __name__ == "__main__":
    main()
