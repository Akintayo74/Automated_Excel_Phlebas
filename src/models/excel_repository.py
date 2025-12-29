import openpyxl
from openpyxl.styles import PatternFill

class ExcelRepository:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = None
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    def load(self):
        try:
            self.wb = openpyxl.load_workbook(self.file_path)
            return True
        except Exception as e:
            print(f"✗ Error loading Excel: {str(e)}")
            return False

    def get_sheet_names(self):
        return self.wb.sheetnames if self.wb else []

    def get_students_from_sheet(self, sheet_name, start_row=3):
        """Yields (row_idx, student_name, current_admission)"""
        ws = self.wb[sheet_name]
        for row_idx in range(start_row, ws.max_row + 1):
            admission_cell = ws.cell(row=row_idx, column=1)
            name_cell = ws.cell(row=row_idx, column=2)
            
            yield {
                'row_idx': row_idx,
                'name': name_cell.value,
                'current_admission': admission_cell.value
            }

    def update_student(self, sheet_name, row_idx, admission_number):
        ws = self.wb[sheet_name]
        admission_cell = ws.cell(row=row_idx, column=1)
        admission_cell.value = admission_number
        admission_cell.fill = self.yellow_fill

    def save(self):
        try:
            # Save to consistency _updated.xlsx
            # If input was already _updated, this overwrites it (Good for single source of truth)
            if '_updated' in self.file_path:
                output_path = self.file_path
            else:
                output_path = self.file_path.replace('.xlsx', '_updated.xlsx')
                
            self.wb.save(output_path)
            return output_path
        except Exception as e:
            print(f"\n✗ Error saving workbook: {str(e)}")
            return None
